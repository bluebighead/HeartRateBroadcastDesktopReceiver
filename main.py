import sys
import os
import asyncio
import time
import json
import csv
import urllib.request
import urllib.error
from datetime import datetime, date
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QMessageBox, QStackedWidget, QTableWidget,
    QTableWidgetItem, QHeaderView, QAbstractItemView, QGraphicsOpacityEffect,
    QDialog, QCheckBox, QLineEdit, QComboBox, QSpinBox, QFormLayout,
    QMenuBar, QAction, QGroupBox, QFileDialog, QTabWidget, QTextEdit,
    QProgressBar
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QTimer, QPropertyAnimation, QEasingCurve, pyqtProperty
from PyQt5.QtGui import QFont
from bleak import BleakClient, BleakScanner
import pyqtgraph as pg

APP_VERSION = "v1.0.2"
GITHUB_REPO = "bluebighead/HeartRateBroadcastDesktopReceiver"
GITHUB_RELEASES_API = f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest"

HEART_RATE_SERVICE_UUID = "0000180D-0000-1000-8000-00805F9B34FB"
HEART_RATE_MEASUREMENT_UUID = "00002A37-0000-1000-8000-00805F9B34FB"


class BleakWorker(QThread):
    heart_rate_received = pyqtSignal(int)
    rr_interval_received = pyqtSignal(list)
    status_changed = pyqtSignal(str)
    error_occurred = pyqtSignal(str)
    connection_lost = pyqtSignal()

    def __init__(self):
        super().__init__()
        self.running = False
        self.client = None
        self.reconnect_count = 0
        self.max_reconnect_attempts = 10
        self.last_heart_rate_time = 0
        self.heart_rate_timeout = 10  # 10秒没有收到心率数据则认为断连

    def run(self):
        self.running = True
        asyncio.run(self._run_ble())

    async def _run_ble(self):
        # 先扫描找到设备
        self.status_changed.emit("正在扫描心率设备...")
        device = await BleakScanner.find_device_by_filter(
            lambda d, ad: HEART_RATE_SERVICE_UUID.lower() in [s.lower() for s in ad.service_uuids]
        )
        
        if not device:
            self.error_occurred.emit("未找到心率设备，请确保设备已开启并靠近电脑")
            return
        
        device_address = device.address
        device_name = device.name or device_address
        self.status_changed.emit(f"已找到设备: {device_name}")
        
        # 循环尝试连接同一个设备
        while self.running:
            try:
                async with BleakClient(device_address) as client:
                    self.client = client
                    self.status_changed.emit(f"已连接: {device_name}")
                    self.reconnect_count = 0  # 重置重连计数器
                    self.last_heart_rate_time = time.time()  # 重置心跳时间
                    
                    await client.start_notify(
                        HEART_RATE_MEASUREMENT_UUID,
                        self._heart_rate_handler
                    )
                    
                    while self.running:
                        # 检查心率数据是否超时
                        if time.time() - self.last_heart_rate_time > self.heart_rate_timeout:
                            raise Exception("心率数据超时，设备可能已断连")
                        await asyncio.sleep(1)  # 每1秒检查一次
                    
                    await client.stop_notify(HEART_RATE_MEASUREMENT_UUID)
                    break  # 正常停止，退出循环
                
            except Exception as e:
                if not self.running:
                    break  # 正常停止，退出循环
                
                self.reconnect_count += 1
                if self.reconnect_count > self.max_reconnect_attempts:
                    self.error_occurred.emit(f"连接失败: 已尝试重连{self.max_reconnect_attempts}次，请检查设备状态")
                    self.connection_lost.emit()
                    break
                
                self.status_changed.emit(f"连接断开，正在尝试重连... ({self.reconnect_count}/{self.max_reconnect_attempts})")
                await asyncio.sleep(2)  # 等待2秒后重试
                
        self.status_changed.emit("已断开连接")
        
    def _heart_rate_handler(self, sender, data):
        # 更新最后收到心率数据的时间
        self.last_heart_rate_time = time.time()
        
        if len(data) < 2:
            return
        
        flags = data[0]
        offset = 1
        
        if flags & 0x01:
            heart_rate = int.from_bytes(data[offset:offset+2], byteorder='little')
            offset += 2
        else:
            heart_rate = data[offset]
            offset += 1
        
        if flags & 0x08:
            offset += 2
        
        rr_intervals = []
        if flags & 0x10:
            while offset + 1 < len(data):
                rr_raw = int.from_bytes(data[offset:offset+2], byteorder='little')
                rr_ms = rr_raw * (1000.0 / 1024.0)
                if rr_ms > 0:
                    rr_intervals.append(rr_ms)
                offset += 2
        
        self.heart_rate_received.emit(heart_rate)
        
        if rr_intervals:
            self.rr_interval_received.emit(rr_intervals)

    def stop(self):
        self.running = False
        self.wait()


class HeartRateRecord:
    def __init__(self, heart_rate, timestamp):
        self.heart_rate = heart_rate
        self.timestamp = timestamp


class CalorieSettingsDialog(QDialog):
    def __init__(self, parent=None, current_settings=None):
        super().__init__(parent)
        self.setWindowTitle("卡路里设置")
        self.setMinimumWidth(350)
        self.settings = current_settings or {
            'enabled': False,
            'weight': 70,
            'age': 30,
            'gender': 'male'
        }
        self.init_ui()
        
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        self.enable_checkbox = QCheckBox("是否开启实时卡路里")
        self.enable_checkbox.setChecked(self.settings['enabled'])
        self.enable_checkbox.stateChanged.connect(self.on_enable_changed)
        self.enable_checkbox.setStyleSheet("font-weight: bold; color: #2C3E50;")
        layout.addWidget(self.enable_checkbox)
        
        layout.addSpacing(10)
        
        form_group = QGroupBox("个人信息")
        form_layout = QFormLayout(form_group)
        
        self.weight_spin = QSpinBox()
        self.weight_spin.setRange(30, 200)
        self.weight_spin.setValue(int(self.settings['weight']))
        self.weight_spin.setSuffix(" kg")
        self.weight_spin.setEnabled(self.settings['enabled'])
        form_layout.addRow("体重:", self.weight_spin)
        
        self.age_spin = QSpinBox()
        self.age_spin.setRange(10, 100)
        self.age_spin.setValue(int(self.settings['age']))
        self.age_spin.setSuffix(" 岁")
        self.age_spin.setEnabled(self.settings['enabled'])
        form_layout.addRow("年龄:", self.age_spin)
        
        self.gender_combo = QComboBox()
        self.gender_combo.addItem("男", "male")
        self.gender_combo.addItem("女", "female")
        if self.settings['gender'] == 'female':
            self.gender_combo.setCurrentIndex(1)
        self.gender_combo.setEnabled(self.settings['enabled'])
        form_layout.addRow("性别:", self.gender_combo)
        
        layout.addWidget(form_group)
        
        layout.addSpacing(20)
        
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        cancel_button = QPushButton("取消")
        cancel_button.setMinimumWidth(80)
        cancel_button.setStyleSheet("""
            QPushButton {
                background-color: #95A5A6;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px;
            }
            QPushButton:hover {
                background-color: #7F8C8D;
            }
        """)
        cancel_button.clicked.connect(self.reject)
        button_layout.addWidget(cancel_button)
        
        self.ok_button = QPushButton("确定")
        self.ok_button.setMinimumWidth(80)
        self.ok_button.setStyleSheet("""
            QPushButton {
                background-color: #27AE60;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px;
            }
            QPushButton:hover {
                background-color: #2ECC71;
            }
        """)
        self.ok_button.clicked.connect(self.on_ok_clicked)
        button_layout.addWidget(self.ok_button)
        
        layout.addLayout(button_layout)
        
        self.on_enable_changed(self.enable_checkbox.checkState())
    
    def on_enable_changed(self, state):
        enabled = state == Qt.Checked
        self.weight_spin.setEnabled(enabled)
        self.age_spin.setEnabled(enabled)
        self.gender_combo.setEnabled(enabled)
        
        if enabled:
            self.weight_spin.setStyleSheet("")
            self.age_spin.setStyleSheet("")
            self.gender_combo.setStyleSheet("")
        else:
            self.weight_spin.setStyleSheet("color: #BDC3C7;")
            self.age_spin.setStyleSheet("color: #BDC3C7;")
            self.gender_combo.setStyleSheet("color: #BDC3C7;")
    
    def on_ok_clicked(self):
        if self.enable_checkbox.isChecked():
            if self.weight_spin.value() <= 0:
                QMessageBox.warning(self, "提示", "请输入有效的体重")
                return
            if self.age_spin.value() <= 0:
                QMessageBox.warning(self, "提示", "请输入有效的年龄")
                return
        
        self.settings = {
            'enabled': self.enable_checkbox.isChecked(),
            'weight': self.weight_spin.value(),
            'age': self.age_spin.value(),
            'gender': self.gender_combo.currentData()
        }
        self.accept()
    
    def get_settings(self):
        return self.settings


class HeartAnimationLabel(QLabel):
    def __init__(self, parent=None):
        super().__init__("❤", parent)
        self.setFont(QFont("Arial", 48))
        self.setStyleSheet("color: #E74C3C;")
        self.setAlignment(Qt.AlignCenter)
        
        self._scale = 1.0
        self.animation = QPropertyAnimation(self, b"scale")
        self.animation.setDuration(300)
        self.animation.setEasingCurve(QEasingCurve.InOutQuad)
        
        self.beat_timer = QTimer()
        self.beat_timer.timeout.connect(self.beat)
        self.is_beating = False
        
    def get_scale(self):
        return self._scale
    
    def set_scale(self, value):
        self._scale = value
        font = self.font()
        base_size = 48
        new_size = int(base_size * value)
        font.setPointSize(new_size)
        self.setFont(font)
    
    scale = pyqtProperty(float, get_scale, set_scale)
    
    def start_beating(self, heart_rate=72):
        if heart_rate > 0:
            interval = int(60000 / heart_rate)
            self.beat_timer.start(max(400, interval))
        self.is_beating = True
    
    def stop_beating(self):
        self.beat_timer.stop()
        self.is_beating = False
        self.set_scale(1.0)
    
    def beat(self):
        self.animation.stop()
        self.animation.setStartValue(1.0)
        self.animation.setKeyValueAt(0.3, 1.3)
        self.animation.setKeyValueAt(0.6, 1.0)
        self.animation.setEndValue(1.0)
        self.animation.start()
    
    def update_heart_rate(self, heart_rate):
        if self.is_beating and heart_rate > 0:
            interval = int(60000 / heart_rate)
            self.beat_timer.setInterval(max(400, interval))


class HomePage(QWidget):
    heart_rate_recorded = pyqtSignal(int, datetime)
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.ble_worker = None
        self.last_heart_rate = None
        self.start_time = None
        self.current_heart_rate = 0
        self.total_calories = 0.0
        self.calorie_settings = {
            'enabled': False,
            'weight': 70,
            'age': 30,
            'gender': 'male'
        }
        self.hrv_enabled = False
        self.obs_enabled = False
        self.obs_file_path = "obs_heart_rate.txt"
        self.heart_rate_timestamps = []
        self.hrv_window_size = 30
        self.real_rr_intervals = []
        self.using_real_rr = False
        self.init_ui()
        
        self.calorie_timer = QTimer()
        self.calorie_timer.timeout.connect(self.update_calories)
        
    def update_time(self):
        """更新实时时间显示"""
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.time_label.setText(current_time)

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setAlignment(Qt.AlignCenter)
        
        # 添加实时时间显示
        self.time_label = QLabel()
        self.time_label.setFont(QFont("Arial", 16))
        self.time_label.setAlignment(Qt.AlignCenter)
        self.time_label.setStyleSheet("color: #2C3E50;")
        layout.addWidget(self.time_label)
        
        # 更新时间的定时器
        self.time_timer = QTimer()
        self.time_timer.timeout.connect(self.update_time)
        self.time_timer.start(1000)  # 每秒更新一次
        self.update_time()  # 初始更新
        
        layout.addSpacing(10)
        
        heart_layout = QHBoxLayout()
        heart_layout.setAlignment(Qt.AlignCenter)
        
        self.heart_animation = HeartAnimationLabel()
        heart_layout.addWidget(self.heart_animation)
        
        self.heart_rate_label = QLabel("--")
        self.heart_rate_label.setFont(QFont("Arial", 72, QFont.Bold))
        self.heart_rate_label.setAlignment(Qt.AlignCenter)
        self.heart_rate_label.setStyleSheet("color: #E74C3C;")
        heart_layout.addWidget(self.heart_rate_label)
        
        layout.addLayout(heart_layout)
        
        bpm_label = QLabel("BPM")
        bpm_label.setFont(QFont("Arial", 18))
        bpm_label.setAlignment(Qt.AlignCenter)
        bpm_label.setStyleSheet("color: #7F8C8D;")
        layout.addWidget(bpm_label)
        
        layout.addSpacing(10)
        
        self.calorie_label = QLabel("消耗卡路里: 0.0 kcal")
        self.calorie_label.setFont(QFont("Arial", 14))
        self.calorie_label.setAlignment(Qt.AlignCenter)
        self.calorie_label.setStyleSheet("color: #F39C12; font-weight: bold;")
        self.calorie_label.hide()
        layout.addWidget(self.calorie_label)
        
        self.duration_label = QLabel("运动时长: 00:00:00")
        self.duration_label.setFont(QFont("Arial", 12))
        self.duration_label.setAlignment(Qt.AlignCenter)
        self.duration_label.setStyleSheet("color: #7F8C8D;")
        layout.addWidget(self.duration_label)
        
        layout.addSpacing(10)
        
        self.hrv_frame = QGroupBox()
        self.hrv_frame.setStyleSheet("""
            QGroupBox {
                background-color: #F8F9FA;
                border: 1px solid #DEE2E6;
                border-radius: 8px;
                margin-top: 5px;
                padding: 10px;
            }
        """)
        hrv_layout = QVBoxLayout(self.hrv_frame)
        hrv_layout.setSpacing(5)
        
        hrv_title_layout = QHBoxLayout()
        hrv_title_layout.setAlignment(Qt.AlignCenter)
        hrv_title_layout.setSpacing(10)
        
        hrv_title = QLabel("心率变异性 (HRV)")
        hrv_title.setFont(QFont("Arial", 11, QFont.Bold))
        hrv_title.setStyleSheet("color: #2C3E50;")
        hrv_title_layout.addWidget(hrv_title)
        
        hrv_help = QLabel("?")
        hrv_help.setFont(QFont("Arial", 12, QFont.Bold))
        hrv_help.setStyleSheet("""
            QLabel {
                color: white;
                background-color: #3498DB;
                border-radius: 10px;
                padding: 0px 6px;
                min-width: 20px;
                min-height: 20px;
            }
        """)
        hrv_help.setFixedSize(20, 20)
        hrv_help.setAlignment(Qt.AlignCenter)
        hrv_help.setToolTip(
            "心率变异性 (HRV) 说明：\n\n"
            "HRV 是衡量自主神经系统活动的重要指标。\n\n"
            "• RMSSD：相邻心跳间隔差值的均方根\n"
            "• 高 HRV：身体恢复良好，压力低\n"
            "• 低 HRV：压力大，需要休息\n\n"
            "数据来源：\n"
            "• (真实RR间期)：设备发送的精确数据\n"
            "• (心率推算)：根据心率估算，仅供参考"
        )
        hrv_title_layout.addWidget(hrv_help)
        
        hrv_layout.addLayout(hrv_title_layout)
        
        self.hrv_value_label = QLabel("RMSSD: -- ms")
        self.hrv_value_label.setFont(QFont("Arial", 12))
        self.hrv_value_label.setAlignment(Qt.AlignCenter)
        self.hrv_value_label.setStyleSheet("color: #3498DB;")
        hrv_layout.addWidget(self.hrv_value_label)
        
        self.hrv_status_label = QLabel("状态: 等待数据...")
        self.hrv_status_label.setFont(QFont("Arial", 11))
        self.hrv_status_label.setAlignment(Qt.AlignCenter)
        self.hrv_status_label.setStyleSheet("color: #7F8C8D;")
        hrv_layout.addWidget(self.hrv_status_label)
        
        self.hrv_bar = QLabel()
        self.hrv_bar.setFixedHeight(8)
        self.hrv_bar.setStyleSheet("background-color: #E0E0E0; border-radius: 4px;")
        hrv_layout.addWidget(self.hrv_bar)
        
        layout.addWidget(self.hrv_frame)
        
        # OBS对接网址显示区域
        obs_url_layout = QHBoxLayout()
        obs_url_layout.setAlignment(Qt.AlignCenter)
        obs_url_layout.setSpacing(10)
        
        self.obs_url_label = QLabel("OBS对接网址:")
        self.obs_url_label.setFont(QFont("Arial", 11))
        self.obs_url_label.setStyleSheet("color: #7F8C8D;")
        
        self.obs_url_display = QLabel("")
        self.obs_url_display.setFont(QFont("Arial", 11))
        self.obs_url_display.setStyleSheet("color: #3498DB; background-color: #F8F9FA; padding: 5px 10px; border: 1px solid #DEE2E6; border-radius: 4px;")
        self.obs_url_display.setMinimumWidth(200)
        
        self.obs_copy_button = QPushButton("复制")
        self.obs_copy_button.setFont(QFont("Arial", 10))
        self.obs_copy_button.setFixedSize(60, 25)
        self.obs_copy_button.setStyleSheet("""
            QPushButton {
                background-color: #3498DB;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 2px;
            }
            QPushButton:hover {
                background-color: #2980B9;
            }
        """)
        self.obs_copy_button.clicked.connect(self.copy_obs_url)
        
        obs_url_layout.addWidget(self.obs_url_label)
        obs_url_layout.addWidget(self.obs_url_display)
        obs_url_layout.addWidget(self.obs_copy_button)
        
        self.obs_url_frame = QWidget()
        self.obs_url_frame.setLayout(obs_url_layout)
        self.obs_url_frame.hide()  # 默认隐藏
        
        layout.addWidget(self.obs_url_frame)
        
        layout.addSpacing(15)
        
        self.status_label = QLabel("未连接")
        self.status_label.setFont(QFont("Arial", 12))
        self.status_label.setAlignment(Qt.AlignCenter)
        self.status_label.setStyleSheet("color: #95A5A6;")
        layout.addWidget(self.status_label)
        
        # 重连状态显示标签
        self.reconnect_status_label = QLabel("")
        self.reconnect_status_label.setFont(QFont("Arial", 10))
        self.reconnect_status_label.setAlignment(Qt.AlignRight)
        self.reconnect_status_label.setStyleSheet("color: #E74C3C; font-weight: bold;")
        layout.addWidget(self.reconnect_status_label)
        
        layout.addSpacing(30)
        
        button_layout = QHBoxLayout()
        
        self.start_button = QPushButton("开始接收")
        self.start_button.setFont(QFont("Arial", 14))
        self.start_button.setMinimumSize(120, 45)
        self.start_button.setStyleSheet("""
            QPushButton {
                background-color: #27AE60;
                color: white;
                border: none;
                border-radius: 8px;
            }
            QPushButton:hover {
                background-color: #2ECC71;
            }
            QPushButton:disabled {
                background-color: #95A5A6;
            }
        """)
        self.start_button.clicked.connect(self.start_receiving)
        button_layout.addWidget(self.start_button)
        
        self.stop_button = QPushButton("停止接收")
        self.stop_button.setFont(QFont("Arial", 14))
        self.stop_button.setMinimumSize(120, 45)
        self.stop_button.setEnabled(False)
        self.stop_button.setStyleSheet("""
            QPushButton {
                background-color: #E74C3C;
                color: white;
                border: none;
                border-radius: 8px;
            }
            QPushButton:hover {
                background-color: #C0392B;
            }
            QPushButton:disabled {
                background-color: #95A5A6;
            }
        """)
        self.stop_button.clicked.connect(self.stop_receiving)
        button_layout.addWidget(self.stop_button)
        
        layout.addLayout(button_layout)
    
    def set_calorie_settings(self, settings):
        self.calorie_settings = settings
        if settings['enabled']:
            self.calorie_label.show()
        else:
            self.calorie_label.hide()
    
    def get_calorie_settings(self):
        return self.calorie_settings

    def start_receiving(self):
        self.start_button.setEnabled(False)
        self.stop_button.setEnabled(True)
        self.heart_rate_label.setText("--")
        self.last_heart_rate = None
        self.start_time = time.time()
        self.total_calories = 0.0
        self.current_heart_rate = 0
        self.calorie_label.setText("消耗卡路里: 0.0 kcal")
        self.duration_label.setText("运动时长: 00:00:00")
        self.heart_rate_timestamps = []
        self.real_rr_intervals = []
        self.using_real_rr = False
        self.hrv_value_label.setText("RMSSD: -- ms")
        self.hrv_status_label.setText("状态: 等待数据...")
        self.hrv_status_label.setStyleSheet("color: #7F8C8D;")
        self.hrv_bar.setStyleSheet("background-color: #E0E0E0; border-radius: 4px;")
        
        self.heart_animation.start_beating()
        self.calorie_timer.start(1000)
        
        self.ble_worker = BleakWorker()
        self.ble_worker.heart_rate_received.connect(self.update_heart_rate)
        self.ble_worker.rr_interval_received.connect(self.update_rr_intervals)
        self.ble_worker.status_changed.connect(self.update_status)
        self.ble_worker.error_occurred.connect(self.show_error)
        self.ble_worker.connection_lost.connect(self.stop_receiving)
        self.ble_worker.start()

    def stop_receiving(self):
        if self.ble_worker:
            self.ble_worker.stop()
            self.ble_worker = None
        
        self.heart_animation.stop_beating()
        self.calorie_timer.stop()
        
        self.start_button.setEnabled(True)
        self.stop_button.setEnabled(False)
        self.heart_rate_label.setText("--")
        self.status_label.setText("未连接")
        self.last_heart_rate = None
        self.current_heart_rate = 0

    def update_heart_rate(self, heart_rate):
        self.heart_rate_label.setText(str(heart_rate))
        self.current_heart_rate = heart_rate
        self.heart_animation.update_heart_rate(heart_rate)
        
        current_time = time.time()
        self.heart_rate_timestamps.append((heart_rate, current_time))
        if len(self.heart_rate_timestamps) > self.hrv_window_size:
            self.heart_rate_timestamps.pop(0)
        
        self.update_hrv()
        self.update_obs_data(heart_rate)
        self.update_obs_url_display()  # 更新OBS对接网址显示
        
        if self.last_heart_rate is None or heart_rate != self.last_heart_rate:
            timestamp = datetime.now()
            self.heart_rate_recorded.emit(heart_rate, timestamp)
            self.last_heart_rate = heart_rate

    def update_obs_data(self, heart_rate):
        """更新OBS数据文件"""
        if self.obs_enabled and hasattr(self, 'obs_file_path') and self.obs_file_path:
            try:
                output_type = getattr(self, 'obs_output_type', 'txt')
                # 确保目录存在
                import os
                os.makedirs(os.path.dirname(os.path.abspath(self.obs_file_path)), exist_ok=True)
                
                if output_type == 'txt':
                    # 文本文件输出
                    with open(self.obs_file_path, 'w') as f:
                        f.write(str(heart_rate))
                elif output_type == 'html':
                    # HTML文件输出
                    image_path = getattr(self, 'obs_image_path', "")
                    
                    # 生成图片HTML
                    if image_path:
                        # 确保图片路径是正确的URL格式
                        import os
                        absolute_path = os.path.abspath(image_path)
                        # 转换为file:// URL格式，确保路径正确
                        file_url = 'file:///' + absolute_path.replace('\\', '/').replace(' ', '%20')
                        image_html = '<img src="{}" alt="Heart">'.format(file_url)
                    else:
                        # 默认心形图标
                        image_html = '<svg viewBox="0 0 24 24" fill="#E74C3C"><path d="M12 21.35l-1.45-1.32C5.4 15.36 2 12.28 2 8.5 2 5.42 4.42 3 7.5 3c1.74 0 3.41.81 4.5 2.09C13.09 3.81 14.76 3 16.5 3 19.58 3 22 5.42 22 8.5c0 3.78-3.4 6.86-8.55 11.54L12 21.35z"></path></svg>'
                    
                    # 构建HTML内容
                    html_content = '<!DOCTYPE html>\n'
                    html_content += '<html>\n'
                    html_content += '<head>\n'
                    html_content += '    <meta charset="UTF-8">\n'
                    html_content += '    <meta name="viewport" content="width=device-width, initial-scale=1.0">\n'
                    html_content += '    <title>心率显示</title>\n'
                    html_content += '    <style>\n'
                    html_content += '        * {\n'
                    html_content += '            box-sizing: border-box;\n'
                    html_content += '        }\n'
                    html_content += '        body {\n'
                    html_content += '            margin: 0;\n'
                    html_content += '            padding: 2vh;\n'
                    html_content += '            font-family: Arial, sans-serif;\n'
                    html_content += '            display: flex;\n'
                    html_content += '            align-items: center;\n'
                    html_content += '            justify-content: center;\n'
                    html_content += '            background-color: transparent;\n'
                    html_content += '            height: 100vh;\n'
                    html_content += '            width: 100vw;\n'
                    html_content += '        }\n'
                    html_content += '        .container {\n'
                    html_content += '            display: flex;\n'
                    html_content += '            align-items: center;\n'
                    html_content += '            max-height: 100%;\n'
                    html_content += '        }\n'
                    html_content += '        .heart-icon {\n'
                    html_content += '            margin-right: 4vw;\n'
                    html_content += '        }\n'
                    html_content += '        .heart-icon img {\n'
                    html_content += '            height: 50vh;\n'
                    html_content += '            width: auto;\n'
                    html_content += '            object-fit: contain;\n'
                    html_content += '        }\n'
                    html_content += '        .heart-icon svg {\n'
                    html_content += '            height: 50vh;\n'
                    html_content += '            width: auto;\n'
                    html_content += '        }\n'
                    html_content += '        .heart-rate {\n'
                    html_content += '            font-size: 50vh;\n'
                    html_content += '            font-weight: bold;\n'
                    html_content += '            color: #E74C3C;\n'
                    html_content += '        }\n'
                    html_content += '        .heart-beat {\n'
                    html_content += '            animation: beat 1s infinite;\n'
                    html_content += '        }\n'
                    html_content += '        @keyframes beat {\n'
                    html_content += '            0% {\n'
                    html_content += '                transform: scale(1);\n'
                    html_content += '            }\n'
                    html_content += '            14% {\n'
                    html_content += '                transform: scale(1.3);\n'
                    html_content += '            }\n'
                    html_content += '            28% {\n'
                    html_content += '                transform: scale(1);\n'
                    html_content += '            }\n'
                    html_content += '            42% {\n'
                    html_content += '                transform: scale(1.3);\n'
                    html_content += '            }\n'
                    html_content += '            70% {\n'
                    html_content += '                transform: scale(1);\n'
                    html_content += '            }\n'
                    html_content += '        }\n'
                    html_content += '    </style>\n'
                    html_content += '    <script>\n'
                    html_content += '        // 使用JavaScript定期刷新页面\n'
                    html_content += '        setInterval(function() {\n'
                    html_content += '            location.reload();\n'
                    html_content += '        }, 1000); // 每秒刷新一次\n'
                    html_content += '    </script>\n'
                    html_content += '</head>\n'
                    html_content += '<body>\n'
                    html_content += '    <div class="container">\n'
                    html_content += '        <div class="heart-icon heart-beat">\n'
                    html_content += '            ' + image_html + '\n'
                    html_content += '        </div>\n'
                    html_content += '        <div class="heart-rate">' + str(heart_rate) + '</div>\n'
                    html_content += '    </div>\n'
                    html_content += '</body>\n'
                    html_content += '</html>\n'
                    
                    # 写入HTML文件
                    with open(self.obs_file_path, 'w', encoding='utf-8') as f:
                        f.write(html_content)
            except Exception as e:
                print(f"OBS文件生成失败: {str(e)}")
                import traceback
                traceback.print_exc()
                from PyQt5.QtWidgets import QMessageBox
                QMessageBox.warning(self, "OBS文件生成失败", f"无法生成OBS文件: {str(e)}")

    def update_calories(self):
        if self.start_time and self.current_heart_rate > 0 and self.calorie_settings['enabled']:
            weight = self.calorie_settings['weight']
            age = self.calorie_settings['age']
            gender = self.calorie_settings['gender']
            hr = self.current_heart_rate
            
            if gender == 'male':
                calories_per_minute = ((-55.0969 + 0.6309 * hr + 0.1988 * weight + 0.2017 * age) / 4.184)
            else:
                calories_per_minute = ((-20.4022 + 0.4472 * hr - 0.1263 * weight + 0.074 * age) / 4.184)
            
            calories_per_second = calories_per_minute / 60
            self.total_calories += calories_per_second
            self.calorie_label.setText(f"消耗卡路里: {self.total_calories:.1f} kcal")
        
        if self.start_time:
            elapsed = time.time() - self.start_time
            hours = int(elapsed // 3600)
            minutes = int((elapsed % 3600) // 60)
            seconds = int(elapsed % 60)
            self.duration_label.setText(f"运动时长: {hours:02d}:{minutes:02d}:{seconds:02d}")

    def update_rr_intervals(self, rr_intervals):
        self.using_real_rr = True
        for rr in rr_intervals:
            self.real_rr_intervals.append(rr)
            if len(self.real_rr_intervals) > self.hrv_window_size:
                self.real_rr_intervals.pop(0)
        self.update_hrv()

    def update_hrv(self):
        if self.using_real_rr and len(self.real_rr_intervals) >= 5:
            rr_intervals = self.real_rr_intervals.copy()
            source_text = "(真实RR间期)"
        elif len(self.heart_rate_timestamps) >= 5:
            rr_intervals = []
            for hr, _ in self.heart_rate_timestamps:
                if hr > 0:
                    rr_interval = 60000.0 / hr
                    rr_intervals.append(rr_interval)
            source_text = "(心率推算)"
        else:
            self.hrv_value_label.setText("RMSSD: -- ms")
            self.hrv_status_label.setText("状态: 等待更多数据...")
            self.hrv_status_label.setStyleSheet("color: #7F8C8D;")
            self.hrv_bar.setStyleSheet("background-color: #E0E0E0; border-radius: 4px;")
            return
        
        if len(rr_intervals) < 5:
            return
        
        squared_diffs = []
        for i in range(1, len(rr_intervals)):
            diff = rr_intervals[i] - rr_intervals[i-1]
            squared_diffs.append(diff ** 2)
        
        if squared_diffs:
            rmssd = (sum(squared_diffs) / len(squared_diffs)) ** 0.5
        else:
            return
        
        self.hrv_value_label.setText(f"RMSSD: {rmssd:.1f} ms {source_text}")
        
        status, color, bar_color = self.get_hrv_status(rmssd)
        self.hrv_status_label.setText(f"状态: {status}")
        self.hrv_status_label.setStyleSheet(f"color: {color};")
        self.hrv_bar.setStyleSheet(f"background-color: {bar_color}; border-radius: 4px;")

    def get_hrv_status(self, rmssd):
        if rmssd >= 100:
            return "优秀 - 恢复良好，压力低", "#27AE60", "#27AE60"
        elif rmssd >= 50:
            return "良好 - 身体状态正常", "#3498DB", "#3498DB"
        elif rmssd >= 20:
            return "一般 - 有一定压力", "#F39C12", "#F39C12"
        else:
            return "较低 - 压力大，建议休息", "#E74C3C", "#E74C3C"

    def update_status(self, status):
        self.status_label.setText(status)
        
        # 更新重连状态显示
        if "重连" in status:
            self.reconnect_status_label.setText(status)
        else:
            self.reconnect_status_label.setText("")

    def show_error(self, message):
        QMessageBox.warning(self, "错误", message)
        self.stop_receiving()

    def update_hrv_visibility(self, visible):
        self.hrv_frame.setVisible(visible)

    def copy_obs_url(self):
        """复制OBS对接网址到剪贴板"""
        from PyQt5.QtWidgets import QApplication, QMessageBox
        clipboard = QApplication.clipboard()
        clipboard.setText(self.obs_url_display.text())
        QMessageBox.information(self, "复制成功", "OBS对接网址已复制到剪贴板")

    def update_obs_url_display(self):
        """更新OBS对接网址显示"""
        if self.obs_enabled and self.obs_file_path:
            import os
            # 生成文件的绝对路径URL
            file_url = f"file:///{os.path.abspath(self.obs_file_path).replace('\\', '/')}"
            self.obs_url_display.setText(file_url)
            self.obs_url_frame.show()
        else:
            self.obs_url_frame.hide()

    def cleanup(self):
        self.stop_receiving()


class StatsWindow(QWidget):
    def __init__(self, record_page, parent=None):
        super().__init__(parent)
        self.record_page = record_page
        self.setWindowTitle("心率统计面板")
        self.setMinimumSize(350, 300)
        self.init_ui()
        
        # 加载保存的体重设置
        self.load_weight_setting()
        
        self.update_timer = QTimer()
        self.update_timer.timeout.connect(self.update_stats)
        self.update_timer.start(500)

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(20)
        
        title_label = QLabel("心率统计")
        title_label.setFont(QFont("Arial", 18, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("color: #2C3E50; margin: 10px;")
        layout.addWidget(title_label)
        
        stats_frame = QGroupBox()
        stats_frame.setStyleSheet("""
            QGroupBox {
                background-color: white;
                border: 2px solid #BDC3C7;
                border-radius: 10px;
                margin-top: 10px;
                padding: 15px;
            }
        """)
        stats_layout = QVBoxLayout(stats_frame)
        stats_layout.setSpacing(15)
        
        self.max_label = QLabel("最大心率: -- BPM")
        self.max_label.setFont(QFont("Arial", 14))
        self.max_label.setStyleSheet("color: #E74C3C;")
        stats_layout.addWidget(self.max_label)
        
        self.min_label = QLabel("最小心率: -- BPM")
        self.min_label.setFont(QFont("Arial", 14))
        self.min_label.setStyleSheet("color: #3498DB;")
        stats_layout.addWidget(self.min_label)
        
        self.avg_label = QLabel("平均心率: -- BPM")
        self.avg_label.setFont(QFont("Arial", 14))
        self.avg_label.setStyleSheet("color: #2ECC71;")
        stats_layout.addWidget(self.avg_label)
        
        self.count_label = QLabel("记录数量: 0")
        self.count_label.setFont(QFont("Arial", 14))
        self.count_label.setStyleSheet("color: #9B59B6;")
        stats_layout.addWidget(self.count_label)
        
        # 体重输入
        weight_layout = QHBoxLayout()
        weight_label = QLabel("体重 (kg):")
        weight_label.setFont(QFont("Arial", 14))
        weight_label.setStyleSheet("color: #3498DB;")
        weight_layout.addWidget(weight_label)
        
        self.weight_input = QLineEdit()
        self.weight_input.setFont(QFont("Arial", 14))
        self.weight_input.setFixedWidth(80)
        self.weight_input.setAlignment(Qt.AlignCenter)
        self.weight_input.setText("70")  # 默认体重
        self.weight_input.textChanged.connect(self.on_weight_changed)
        weight_layout.addWidget(self.weight_input)
        
        stats_layout.addLayout(weight_layout)
        
        # 卡路里总消耗
        self.calories_label = QLabel("卡路里总消耗: 0.0 kcal")
        self.calories_label.setFont(QFont("Arial", 14))
        self.calories_label.setStyleSheet("color: #F39C12;")
        stats_layout.addWidget(self.calories_label)
        
        layout.addWidget(stats_frame)
        
        self.range_label = QLabel("心率范围: --")
        self.range_label.setFont(QFont("Arial", 12))
        self.range_label.setAlignment(Qt.AlignCenter)
        self.range_label.setStyleSheet("color: #7F8C8D; margin: 10px;")
        layout.addWidget(self.range_label)

    def update_stats(self):
        records = self.record_page.records
        if not records:
            self.max_label.setText("最大心率: -- BPM")
            self.min_label.setText("最小心率: -- BPM")
            self.avg_label.setText("平均心率: -- BPM")
            self.count_label.setText("记录数量: 0")
            self.calories_label.setText("卡路里总消耗: 0.0 kcal")
            self.range_label.setText("心率范围: --")
            return
        
        heart_rates = [r.heart_rate for r in records]
        max_hr = max(heart_rates)
        min_hr = min(heart_rates)
        avg_hr = sum(heart_rates) / len(heart_rates)
        
        # 计算卡路里总消耗
        total_calories = 0.0
        # 假设每记录间隔为1分钟，使用简化的卡路里计算公式
        # 卡路里 = 0.014 * 体重(kg) * 心率 * 时间(分钟)
        # 获取用户输入的体重
        try:
            weight = float(self.weight_input.text())
        except ValueError:
            weight = 70  # 默认体重
        for i in range(len(records) - 1):
            current_hr = records[i].heart_rate
            next_hr = records[i+1].heart_rate
            # 计算平均心率
            avg_hr_interval = (current_hr + next_hr) / 2
            # 计算时间差（分钟）
            time_diff = (records[i+1].timestamp - records[i].timestamp).total_seconds() / 60
            # 计算卡路里消耗
            calories = 0.014 * weight * avg_hr_interval * time_diff
            total_calories += calories
        
        self.max_label.setText(f"最大心率: {max_hr} BPM")
        self.min_label.setText(f"最小心率: {min_hr} BPM")
        self.avg_label.setText(f"平均心率: {avg_hr:.1f} BPM")
        self.count_label.setText(f"记录数量: {len(records)}")
        self.calories_label.setText(f"卡路里总消耗: {total_calories:.1f} kcal")
        self.range_label.setText(f"心率范围: {min_hr} - {max_hr} BPM (波动 {max_hr - min_hr} BPM)")

    def on_weight_changed(self):
        """处理体重输入变化"""
        # 保存体重设置
        self.save_weight_setting()
        # 更新统计信息
        self.update_stats()
    
    def load_weight_setting(self):
        """加载保存的体重设置"""
        import os
        import json
        
        # 保存设置的文件路径
        settings_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "settings.json")
        
        try:
            if os.path.exists(settings_file):
                with open(settings_file, 'r', encoding='utf-8') as f:
                    settings = json.load(f)
                    if 'weight' in settings:
                        self.weight_input.setText(str(settings['weight']))
        except Exception as e:
            print(f"加载体重设置失败: {str(e)}")
    
    def save_weight_setting(self):
        """保存体重设置"""
        import os
        import json
        
        # 保存设置的文件路径
        settings_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "settings.json")
        
        try:
            # 获取当前体重输入
            weight_text = self.weight_input.text()
            if weight_text and weight_text.replace('.', '').isdigit():
                weight = float(weight_text)
                
                # 读取现有设置
                settings = {}
                if os.path.exists(settings_file):
                    with open(settings_file, 'r', encoding='utf-8') as f:
                        settings = json.load(f)
                
                # 更新体重设置
                settings['weight'] = weight
                
                # 保存设置
                with open(settings_file, 'w', encoding='utf-8') as f:
                    json.dump(settings, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"保存体重设置失败: {str(e)}")

    def closeEvent(self, event):
        self.update_timer.stop()
        event.accept()


class ChartWindow(QWidget):
    def __init__(self, record_page, parent=None):
        super().__init__(parent)
        self.record_page = record_page
        self.setWindowTitle("心率数据可视化")
        self.setMinimumSize(800, 600)
        self.init_ui()
        
        self.update_timer = QTimer()
        self.update_timer.timeout.connect(self.update_charts)
        self.update_timer.start(500)

    def init_ui(self):
        layout = QVBoxLayout(self)
        
        pg.setConfigOptions(antialias=True)
        
        # 添加选项卡
        self.tab_widget = QTabWidget()
        layout.addWidget(self.tab_widget)
        
        # 折线图标签页
        self.line_chart_tab = QWidget()
        self.tab_widget.addTab(self.line_chart_tab, "折线图")
        self.init_line_chart()
        
        # 热图标签页
        self.heatmap_tab = QWidget()
        self.tab_widget.addTab(self.heatmap_tab, "心率热图")
        self.init_heatmap()
        
        # 趋势图标签页
        self.trend_tab = QWidget()
        self.tab_widget.addTab(self.trend_tab, "周/月趋势")
        self.init_trend_chart()
        
        info_layout = QHBoxLayout()
        
        self.current_label = QLabel("当前心率: -- BPM")
        self.current_label.setFont(QFont("Arial", 12))
        self.current_label.setStyleSheet("color: #2C3E50;")
        info_layout.addWidget(self.current_label)
        
        self.count_label = QLabel("数据点: 0")
        self.count_label.setFont(QFont("Arial", 12))
        self.count_label.setStyleSheet("color: #2C3E50;")
        info_layout.addWidget(self.count_label)
        
        info_layout.addStretch()
        layout.addLayout(info_layout)

    def init_line_chart(self):
        layout = QVBoxLayout(self.line_chart_tab)
        
        self.plot_widget = pg.PlotWidget()
        self.plot_widget.setBackground('w')
        self.plot_widget.setTitle("心率变化趋势", color='#2C3E50', size='14pt')
        self.plot_widget.setLabel('left', '心率', 'BPM', color='#2C3E50')
        self.plot_widget.setLabel('bottom', '序号', units='', color='#2C3E50')
        self.plot_widget.showGrid(x=True, y=True, alpha=0.3)
        self.plot_widget.setYRange(40, 200)
        
        self.curve = self.plot_widget.plot(
            pen=pg.mkPen(color='#E74C3C', width=2),
            symbol='o',
            symbolSize=8,
            symbolBrush='#E74C3C'
        )
        
        layout.addWidget(self.plot_widget)

    def init_heatmap(self):
        layout = QVBoxLayout(self.heatmap_tab)
        
        self.heatmap_widget = pg.PlotWidget()
        self.heatmap_widget.setBackground('w')
        self.heatmap_widget.setTitle("心率热图", color='#2C3E50', size='14pt')
        self.heatmap_widget.setLabel('left', '小时', units='', color='#2C3E50')
        self.heatmap_widget.setLabel('bottom', '日期', units='', color='#2C3E50')
        
        layout.addWidget(self.heatmap_widget)
        
        # 添加颜色说明
        color_layout = QHBoxLayout()
        color_layout.setAlignment(Qt.AlignCenter)
        
        # 低心率（蓝色）
        low_label = QLabel("低心率")
        low_label.setFont(QFont("Arial", 10))
        low_label.setStyleSheet("color: #2C3E50;")
        color_layout.addWidget(low_label)
        
        low_color = QLabel()
        low_color.setFixedSize(20, 20)
        low_color.setStyleSheet("background-color: #0000FF;")
        color_layout.addWidget(low_color)
        
        # 中心率（紫色）
        mid_label = QLabel("中心率")
        mid_label.setFont(QFont("Arial", 10))
        mid_label.setStyleSheet("color: #2C3E50;")
        color_layout.addWidget(mid_label)
        
        mid_color = QLabel()
        mid_color.setFixedSize(20, 20)
        mid_color.setStyleSheet("background-color: #800080;")
        color_layout.addWidget(mid_color)
        
        # 高心率（红色）
        high_label = QLabel("高心率")
        high_label.setFont(QFont("Arial", 10))
        high_label.setStyleSheet("color: #2C3E50;")
        color_layout.addWidget(high_label)
        
        high_color = QLabel()
        high_color.setFixedSize(20, 20)
        high_color.setStyleSheet("background-color: #FF0000;")
        color_layout.addWidget(high_color)
        
        layout.addLayout(color_layout)

    def init_trend_chart(self):
        layout = QVBoxLayout(self.trend_tab)
        
        # 添加周/月选择按钮
        button_layout = QHBoxLayout()
        
        self.week_button = QPushButton("周趋势")
        self.week_button.setFont(QFont("Arial", 10))
        self.week_button.setCheckable(True)
        self.week_button.setChecked(True)
        self.week_button.clicked.connect(lambda: self.set_trend_mode('week'))
        button_layout.addWidget(self.week_button)
        
        self.month_button = QPushButton("月趋势")
        self.month_button.setFont(QFont("Arial", 10))
        self.month_button.setCheckable(True)
        self.month_button.clicked.connect(lambda: self.set_trend_mode('month'))
        button_layout.addWidget(self.month_button)
        
        button_layout.addStretch()
        layout.addLayout(button_layout)
        
        self.trend_widget = pg.PlotWidget()
        self.trend_widget.setBackground('w')
        self.trend_widget.setTitle("周心率趋势", color='#2C3E50', size='14pt')
        self.trend_widget.setLabel('left', '平均心率', 'BPM', color='#2C3E50')
        self.trend_widget.setLabel('bottom', '日期', units='', color='#2C3E50')
        self.trend_widget.showGrid(x=True, y=True, alpha=0.3)
        self.trend_widget.setYRange(40, 200)
        
        self.trend_curve = self.trend_widget.plot(
            pen=pg.mkPen(color='#3498DB', width=2),
            symbol='o',
            symbolSize=8,
            symbolBrush='#3498DB'
        )
        
        layout.addWidget(self.trend_widget)
        
        self.trend_mode = 'week'

    def set_trend_mode(self, mode):
        self.trend_mode = mode
        self.week_button.setChecked(mode == 'week')
        self.month_button.setChecked(mode == 'month')
        
        if mode == 'week':
            self.trend_widget.setTitle("周心率趋势", color='#2C3E50', size='14pt')
        else:
            self.trend_widget.setTitle("月心率趋势", color='#2C3E50', size='14pt')
        
        self.update_charts()

    def update_charts(self):
        records = self.record_page.records
        if not records:
            # 更新折线图
            self.curve.setData([], [])
            
            # 清空热图
            self.heatmap_widget.clear()
            
            # 更新趋势图
            self.trend_curve.setData([], [])
            
            self.current_label.setText("当前心率: -- BPM")
            self.count_label.setText("数据点: 0")
            return
        
        # 更新基本信息
        y_data = [r.heart_rate for r in records]
        if y_data:
            self.current_label.setText(f"当前心率: {y_data[-1]} BPM")
        self.count_label.setText(f"数据点: {len(records)}")
        
        # 更新折线图
        self.update_line_chart(records)
        
        # 更新热图
        self.update_heatmap(records)
        
        # 更新趋势图
        self.update_trend_chart(records)

    def update_line_chart(self, records):
        x_data = list(range(1, len(records) + 1))
        y_data = [r.heart_rate for r in records]
        
        self.curve.setData(x_data, y_data)

    def update_heatmap(self, records):
        # 准备热图数据
        from collections import defaultdict
        import numpy as np
        
        # 按日期和小时分组
        data = defaultdict(lambda: defaultdict(int))
        dates = set()
        hours = set()
        
        for record in records:
            date = record.timestamp.date()
            hour = record.timestamp.hour
            data[date][hour] = record.heart_rate
            dates.add(date)
            hours.add(hour)
        
        if not dates:
            return
        
        # 排序日期和小时
        sorted_dates = sorted(dates)
        sorted_hours = sorted(hours)
        
        # 清除现有热图
        self.heatmap_widget.clear()
        
        # 使用散点图模拟热图效果
        scatter_data = []
        colors = []
        
        for i, date in enumerate(sorted_dates):
            for j, hour in enumerate(sorted_hours):
                if hour in data[date]:
                    heart_rate = data[date][hour]
                    # 计算颜色（基于心率值）
                    # 心率范围映射到0-255
                    color_val = int((heart_rate - 40) / (200 - 40) * 255)
                    color_val = max(0, min(255, color_val))
                    # 使用蓝色到红色的渐变
                    color = pg.mkColor((color_val, 0, 255 - color_val))
                    scatter_data.append({'pos': (i, j), 'size': 20, 'pen': pg.mkPen(None), 'brush': color})
        
        # 创建散点图
        scatter = pg.ScatterPlotItem()
        scatter.addPoints(**{k: [d[k] for d in scatter_data] for k in scatter_data[0]}) if scatter_data else None
        
        # 添加散点图到 plot widget
        if scatter_data:
            self.heatmap_widget.addItem(scatter)
        
        # 设置轴标签
        self.heatmap_widget.setXRange(-0.5, len(sorted_dates) - 0.5)
        self.heatmap_widget.setYRange(-0.5, len(sorted_hours) - 0.5)
        
        # 设置刻度标签
        self.heatmap_widget.getAxis('bottom').setTicks([[(i, str(date)) for i, date in enumerate(sorted_dates)]])
        self.heatmap_widget.getAxis('left').setTicks([[(i, str(hour)) for i, hour in enumerate(sorted_hours)]])

    def update_trend_chart(self, records):
        # 从存储的历史数据文件中读取数据
        from collections import defaultdict
        import os
        import json
        from datetime import datetime, date
        
        daily_data = defaultdict(list)
        
        # 首先处理当前内存中的记录
        for record in records:
            record_date = record.timestamp.date()
            daily_data[record_date].append(record.heart_rate)
        
        # 然后读取存储的历史数据文件
        if hasattr(self, 'data_dir') and os.path.exists(self.data_dir):
            # 遍历数据目录中的所有JSON文件
            for filename in os.listdir(self.data_dir):
                if filename.startswith('heart_rate_') and filename.endswith('.json'):
                    file_path = os.path.join(self.data_dir, filename)
                    try:
                        with open(file_path, 'r', encoding='utf-8') as f:
                            data = json.load(f)
                        
                        # 处理记录数据
                        for record_data in data.get('records', []):
                            if isinstance(record_data, list) and len(record_data) == 2:
                                # 新格式：[时间戳, 心率值]
                                timestamp = datetime.fromtimestamp(record_data[0])
                                heart_rate = record_data[1]
                                record_date = timestamp.date()
                                daily_data[record_date].append(heart_rate)
                            elif isinstance(record_data, dict):
                                # 旧格式
                                heart_rate = record_data.get('heart_rate')
                                timestamp_str = record_data.get('timestamp')
                                if heart_rate and timestamp_str:
                                    timestamp = datetime.fromisoformat(timestamp_str)
                                    record_date = timestamp.date()
                                    daily_data[record_date].append(heart_rate)
                    except Exception as e:
                        print(f"读取历史数据文件失败 {filename}: {str(e)}")
        
        # 计算每天的平均心率
        date_avg = []
        for record_date, rates in daily_data.items():
            if rates:
                avg_rate = sum(rates) / len(rates)
                date_avg.append((record_date, avg_rate))
        
        # 按日期排序
        date_avg.sort(key=lambda x: x[0])
        
        # 根据趋势模式筛选数据
        if self.trend_mode == 'week':
            # 只显示最近7天
            if len(date_avg) > 7:
                date_avg = date_avg[-7:]
        else:
            # 只显示最近30天
            if len(date_avg) > 30:
                date_avg = date_avg[-30:]
        
        # 准备数据
        x_data = list(range(len(date_avg)))
        y_data = [avg for _, avg in date_avg]
        
        # 更新趋势图
        self.trend_curve.setData(x_data, y_data)
        
        # 设置X轴标签
        dates = [str(record_date) for record_date, _ in date_avg]
        self.trend_widget.getAxis('bottom').setTicks([[(i, date) for i, date in enumerate(dates)]])

    def closeEvent(self, event):
        self.update_timer.stop()
        event.accept()


class SleepAnalysisWindow(QWidget):
    def __init__(self, record_page, parent=None):
        super().__init__(parent)
        self.record_page = record_page
        self.setWindowTitle("睡眠质量分析")
        self.setMinimumSize(800, 600)
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # 标题
        title_label = QLabel("睡眠质量分析")
        title_label.setFont(QFont("Arial", 18, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("color: #2C3E50; margin: 10px;")
        layout.addWidget(title_label)
        
        # 睡眠分析结果
        analysis_group = QGroupBox("睡眠分析结果")
        analysis_layout = QVBoxLayout(analysis_group)
        
        # 睡眠状态分布
        self.sleep_distribution_label = QLabel("睡眠状态分布：")
        self.sleep_distribution_label.setFont(QFont("Arial", 12))
        analysis_layout.addWidget(self.sleep_distribution_label)
        
        # 睡眠质量评分
        self.sleep_quality_label = QLabel("睡眠质量评分：")
        self.sleep_quality_label.setFont(QFont("Arial", 12))
        analysis_layout.addWidget(self.sleep_quality_label)
        
        # 睡眠时长
        self.sleep_duration_label = QLabel("睡眠时长：")
        self.sleep_duration_label.setFont(QFont("Arial", 12))
        analysis_layout.addWidget(self.sleep_duration_label)
        
        # 入睡时间
        self.sleep_start_label = QLabel("入睡时间：")
        self.sleep_start_label.setFont(QFont("Arial", 12))
        analysis_layout.addWidget(self.sleep_start_label)
        
        # 清醒时间
        self.sleep_end_label = QLabel("清醒时间：")
        self.sleep_end_label.setFont(QFont("Arial", 12))
        analysis_layout.addWidget(self.sleep_end_label)
        
        # 深睡比例
        self.deep_sleep_label = QLabel("深睡比例：")
        self.deep_sleep_label.setFont(QFont("Arial", 12))
        analysis_layout.addWidget(self.deep_sleep_label)
        
        layout.addWidget(analysis_group)
        
        # 睡眠状态图表
        chart_group = QGroupBox("睡眠状态趋势")
        chart_layout = QVBoxLayout(chart_group)
        
        self.chart_widget = pg.PlotWidget()
        self.chart_widget.setBackground('w')
        self.chart_widget.setTitle("睡眠状态变化", color='#2C3E50', size='14pt')
        self.chart_widget.setLabel('left', '睡眠状态', units='', color='#2C3E50')
        self.chart_widget.setLabel('bottom', '时间', units='', color='#2C3E50')
        self.chart_widget.showGrid(x=True, y=True, alpha=0.3)
        self.chart_widget.setYRange(-0.5, 2.5)
        
        # 设置Y轴刻度
        self.chart_widget.getAxis('left').setTicks([[(0, '清醒'), (1, '浅睡'), (2, '深睡')]])
        
        self.sleep_curve = self.chart_widget.plot(
            pen=pg.mkPen(color='#27AE60', width=2),
            symbol='o',
            symbolSize=8,
            symbolBrush='#27AE60'
        )
        
        chart_layout.addWidget(self.chart_widget)
        layout.addWidget(chart_group)
        
        # 睡眠状态分布
        pie_chart_group = QGroupBox("睡眠状态分布")
        pie_chart_layout = QVBoxLayout(pie_chart_group)
        
        # 使用水平进度条显示睡眠状态分布
        self.sleep_distribution_layout = QVBoxLayout()
        
        # 创建三个进度条，分别对应清醒、浅睡、深睡
        self.awake_bar = QProgressBar()
        self.awake_bar.setStyleSheet("QProgressBar::chunk { background-color: #E74C3C; }")
        self.awake_bar.setFormat("清醒: 0%")
        
        self.light_sleep_bar = QProgressBar()
        self.light_sleep_bar.setStyleSheet("QProgressBar::chunk { background-color: #F39C12; }")
        self.light_sleep_bar.setFormat("浅睡: 0%")
        
        self.deep_sleep_bar = QProgressBar()
        self.deep_sleep_bar.setStyleSheet("QProgressBar::chunk { background-color: #27AE60; }")
        self.deep_sleep_bar.setFormat("深睡: 0%")
        
        self.sleep_distribution_layout.addWidget(self.awake_bar)
        self.sleep_distribution_layout.addWidget(self.light_sleep_bar)
        self.sleep_distribution_layout.addWidget(self.deep_sleep_bar)
        
        pie_chart_layout.addLayout(self.sleep_distribution_layout)
        
        layout.addWidget(pie_chart_group)
        
        # 改善建议
        advice_group = QGroupBox("改善建议")
        advice_layout = QVBoxLayout(advice_group)
        
        self.advice_text = QTextEdit()
        self.advice_text.setReadOnly(True)
        self.advice_text.setStyleSheet("background-color: #F8F9FA; border: 1px solid #E9ECEF;")
        advice_layout.addWidget(self.advice_text)
        
        layout.addWidget(advice_group)
        
        # 按钮布局
        button_layout = QHBoxLayout()
        
        analyze_button = QPushButton("分析睡眠")
        analyze_button.setFont(QFont("Arial", 12))
        analyze_button.setMinimumSize(120, 35)
        analyze_button.setStyleSheet("""
            QPushButton {
                background-color: #3498DB;
                color: white;
                border: none;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #2980B9;
            }
        """)
        analyze_button.clicked.connect(self.analyze_sleep)
        button_layout.addWidget(analyze_button)
        
        export_button = QPushButton("导出报告")
        export_button.setFont(QFont("Arial", 12))
        export_button.setMinimumSize(120, 35)
        export_button.setStyleSheet("""
            QPushButton {
                background-color: #9B59B6;
                color: white;
                border: none;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #8E44AD;
            }
        """)
        export_button.clicked.connect(self.export_report)
        button_layout.addWidget(export_button)
        
        button_layout.addStretch()
        layout.addLayout(button_layout)
        
        # 初始分析
        self.analyze_sleep()

    def analyze_sleep(self):
        records = self.record_page.records
        if not records:
            QMessageBox.warning(self, "提示", "没有足够的心率数据进行睡眠分析")
            return
        
        # 智能睡眠检测：基于心率模式自动识别睡眠时段
        sleep_periods = self.detect_sleep_periods(records)
        
        if not sleep_periods:
            QMessageBox.warning(self, "提示", "未检测到睡眠时段")
            return
        
        # 合并所有睡眠时段的记录
        sleep_records = []
        for start_idx, end_idx in sleep_periods:
            sleep_records.extend(records[start_idx:end_idx+1])
        
        # 识别睡眠状态
        sleep_states = self.identify_sleep_states(sleep_records)
        
        # 计算睡眠质量指标
        sleep_quality = self.calculate_sleep_quality(sleep_states, sleep_records)
        
        # 更新UI
        self.update_ui(sleep_states, sleep_quality, sleep_records)

    def detect_sleep_periods(self, records):
        """
        基于心率模式自动检测睡眠时段
        算法：
        1. 计算心率移动平均值
        2. 识别持续的低心率时段（可能的睡眠时段）
        3. 过滤掉过短的时段
        """
        if len(records) < 10:  # 数据量不足
            return []
        
        # 计算移动平均心率（窗口大小为5）
        window_size = 5
        moving_averages = []
        
        for i in range(len(records)):
            start = max(0, i - window_size + 1)
            window = records[start:i+1]
            avg_hr = sum(r.heart_rate for r in window) / len(window)
            moving_averages.append(avg_hr)
        
        # 识别睡眠时段
        sleep_periods = []
        current_period = []
        sleep_threshold = 65  # 睡眠心率阈值
        min_sleep_duration = 10  # 最小睡眠时长（记录数）
        
        for i, avg_hr in enumerate(moving_averages):
            if avg_hr < sleep_threshold:
                current_period.append(i)
            else:
                if len(current_period) >= min_sleep_duration:
                    # 记录睡眠时段的起始和结束索引
                    sleep_periods.append((current_period[0], current_period[-1]))
                current_period = []
        
        # 处理最后一个时段
        if len(current_period) >= min_sleep_duration:
            sleep_periods.append((current_period[0], current_period[-1]))
        
        return sleep_periods

    def identify_sleep_states(self, records):
        """
        识别睡眠状态：
        - 清醒：心率 > 70 BPM
        - 浅睡：60-70 BPM
        - 深睡：< 60 BPM
        """
        states = []
        for record in records:
            heart_rate = record.heart_rate
            if heart_rate > 70:
                states.append(0)  # 清醒
            elif 60 <= heart_rate <= 70:
                states.append(1)  # 浅睡
            else:
                states.append(2)  # 深睡
        return states

    def calculate_sleep_quality(self, states, records):
        """
        计算睡眠质量：
        - 睡眠时长
        - 深睡比例
        - 睡眠质量评分
        """
        # 计算睡眠时长（分钟）
        start_time = records[0].timestamp
        end_time = records[-1].timestamp
        duration_minutes = (end_time - start_time).total_seconds() / 60
        
        # 计算深睡比例
        deep_sleep_count = states.count(2)
        total_sleep_count = len(states)
        deep_sleep_ratio = (deep_sleep_count / total_sleep_count) * 100 if total_sleep_count > 0 else 0
        
        # 计算睡眠质量评分（0-100）
        # 基于睡眠时长和深睡比例
        duration_score = min(100, (duration_minutes / 480) * 100)  # 8小时为满分
        deep_sleep_score = min(100, deep_sleep_ratio * 2)  # 50%深睡为满分
        quality_score = (duration_score * 0.6) + (deep_sleep_score * 0.4)
        
        return {
            'duration_minutes': duration_minutes,
            'deep_sleep_ratio': deep_sleep_ratio,
            'quality_score': quality_score,
            'states': states
        }

    def update_ui(self, states, sleep_quality, records):
        # 更新睡眠状态分布
        awake_count = states.count(0)
        light_sleep_count = states.count(1)
        deep_sleep_count = states.count(2)
        total_count = len(states)
        
        distribution_text = f"清醒: {awake_count/total_count*100:.1f}%, 浅睡: {light_sleep_count/total_count*100:.1f}%, 深睡: {deep_sleep_count/total_count*100:.1f}%"
        self.sleep_distribution_label.setText(f"睡眠状态分布：{distribution_text}")
        
        # 更新睡眠质量评分
        quality_score = sleep_quality['quality_score']
        quality_level = "优秀" if quality_score >= 80 else "良好" if quality_score >= 60 else "一般" if quality_score >= 40 else "较差"
        self.sleep_quality_label.setText(f"睡眠质量评分：{quality_score:.1f} ({quality_level})")
        
        # 更新睡眠时长
        duration_hours = sleep_quality['duration_minutes'] / 60
        self.sleep_duration_label.setText(f"睡眠时长：{duration_hours:.1f} 小时")
        
        # 更新入睡时间和清醒时间
        if records:
            sleep_start_time = records[0].timestamp.strftime("%Y-%m-%d %H:%M:%S")
            sleep_end_time = records[-1].timestamp.strftime("%Y-%m-%d %H:%M:%S")
            self.sleep_start_label.setText(f"入睡时间：{sleep_start_time}")
            self.sleep_end_label.setText(f"清醒时间：{sleep_end_time}")
        
        # 更新深睡比例
        self.deep_sleep_label.setText(f"深睡比例：{sleep_quality['deep_sleep_ratio']:.1f}%")
        
        # 更新睡眠状态图表
        x_data = list(range(len(records)))
        y_data = states
        self.sleep_curve.setData(x_data, y_data)
        
        # 更新睡眠状态分布饼图
        self.update_pie_chart(states)
        
        # 更新改善建议
        self.update_advice(sleep_quality)

    def update_pie_chart(self, states):
        """更新睡眠状态分布"""
        # 计算各状态的数量
        awake_count = states.count(0)
        light_sleep_count = states.count(1)
        deep_sleep_count = states.count(2)
        total_count = len(states)
        
        if total_count == 0:
            # 重置进度条
            self.awake_bar.setValue(0)
            self.awake_bar.setFormat("清醒: 0%")
            self.light_sleep_bar.setValue(0)
            self.light_sleep_bar.setFormat("浅睡: 0%")
            self.deep_sleep_bar.setValue(0)
            self.deep_sleep_bar.setFormat("深睡: 0%")
            return
        
        # 计算百分比
        awake_percent = awake_count / total_count * 100
        light_sleep_percent = light_sleep_count / total_count * 100
        deep_sleep_percent = deep_sleep_count / total_count * 100
        
        # 更新进度条
        self.awake_bar.setValue(int(awake_percent))
        self.awake_bar.setFormat(f"清醒: {awake_percent:.1f}%")
        
        self.light_sleep_bar.setValue(int(light_sleep_percent))
        self.light_sleep_bar.setFormat(f"浅睡: {light_sleep_percent:.1f}%")
        
        self.deep_sleep_bar.setValue(int(deep_sleep_percent))
        self.deep_sleep_bar.setFormat(f"深睡: {deep_sleep_percent:.1f}%")

    def update_advice(self, sleep_quality):
        advice = []
        quality_score = sleep_quality['quality_score']
        duration_hours = sleep_quality['duration_minutes'] / 60
        deep_sleep_ratio = sleep_quality['deep_sleep_ratio']
        
        if duration_hours < 7:
            advice.append("• 建议增加睡眠时间，理想睡眠时长为7-8小时")
        elif duration_hours > 9:
            advice.append("• 睡眠时间过长，建议控制在7-8小时")
        
        if deep_sleep_ratio < 20:
            advice.append("• 深睡比例偏低，建议保持规律的作息时间")
            advice.append("• 睡前避免使用电子设备，创造安静的睡眠环境")
        
        if quality_score < 60:
            advice.append("• 睡眠质量较差，建议：")
            advice.append("  - 睡前进行放松活动，如冥想或深呼吸")
            advice.append("  - 保持卧室温度适宜（18-20°C）")
            advice.append("  - 避免睡前饮用含咖啡因的饮料")
        elif quality_score < 80:
            advice.append("• 睡眠质量良好，可进一步改善：")
            advice.append("  - 保持规律的起床和睡觉时间")
            advice.append("  - 睡前避免剧烈运动")
        else:
            advice.append("• 睡眠质量优秀，继续保持良好的睡眠习惯！")
        
        self.advice_text.setText("\n".join(advice))

    def export_report(self):
        records = self.record_page.records
        if not records:
            QMessageBox.warning(self, "提示", "没有足够的心率数据生成报告")
            return
        
        # 智能睡眠检测：基于心率模式自动识别睡眠时段
        sleep_periods = self.detect_sleep_periods(records)
        
        if not sleep_periods:
            QMessageBox.warning(self, "提示", "未检测到睡眠时段")
            return
        
        # 合并所有睡眠时段的记录
        sleep_records = []
        for start_idx, end_idx in sleep_periods:
            sleep_records.extend(records[start_idx:end_idx+1])
        
        # 识别睡眠状态
        sleep_states = self.identify_sleep_states(sleep_records)
        
        # 计算睡眠质量指标
        sleep_quality = self.calculate_sleep_quality(sleep_states, sleep_records)
        
        # 生成报告
        report = self.generate_report(sleep_quality, sleep_records, sleep_states)
        
        # 保存报告
        file_path, _ = QFileDialog.getSaveFileName(self, "保存睡眠报告", "sleep_report.txt", "文本文件 (*.txt)")
        if file_path:
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(report)
            QMessageBox.information(self, "成功", f"睡眠报告已导出到: {file_path}")

    def generate_report(self, sleep_quality, records, states):
        start_time = records[0].timestamp
        end_time = records[-1].timestamp
        
        report = f"睡眠质量分析报告\n"
        report += f"=" * 50 + "\n"
        report += f"分析时间范围: {start_time.strftime('%Y-%m-%d %H:%M:%S')} 至 {end_time.strftime('%Y-%m-%d %H:%M:%S')}\n"
        report += f"睡眠时长: {sleep_quality['duration_minutes']/60:.1f} 小时\n"
        report += f"睡眠质量评分: {sleep_quality['quality_score']:.1f}\n"
        report += f"深睡比例: {sleep_quality['deep_sleep_ratio']:.1f}%\n"
        
        # 睡眠状态分布
        awake_count = states.count(0)
        light_sleep_count = states.count(1)
        deep_sleep_count = states.count(2)
        total_count = len(states)
        
        report += "\n睡眠状态分布:\n"
        report += f"  清醒: {awake_count/total_count*100:.1f}% ({awake_count} 次记录)\n"
        report += f"  浅睡: {light_sleep_count/total_count*100:.1f}% ({light_sleep_count} 次记录)\n"
        report += f"  深睡: {deep_sleep_count/total_count*100:.1f}% ({deep_sleep_count} 次记录)\n"
        
        # 改善建议
        report += "\n改善建议:\n"
        if sleep_quality['duration_minutes']/60 < 7:
            report += "• 建议增加睡眠时间，理想睡眠时长为7-8小时\n"
        elif sleep_quality['duration_minutes']/60 > 9:
            report += "• 睡眠时间过长，建议控制在7-8小时\n"
        
        if sleep_quality['deep_sleep_ratio'] < 20:
            report += "• 深睡比例偏低，建议保持规律的作息时间\n"
            report += "• 睡前避免使用电子设备，创造安静的睡眠环境\n"
        
        if sleep_quality['quality_score'] < 60:
            report += "• 睡眠质量较差，建议：\n"
            report += "  - 睡前进行放松活动，如冥想或深呼吸\n"
            report += "  - 保持卧室温度适宜（18-20°C）\n"
            report += "  - 避免睡前饮用含咖啡因的饮料\n"
        elif sleep_quality['quality_score'] < 80:
            report += "• 睡眠质量良好，可进一步改善：\n"
            report += "  - 保持规律的起床和睡觉时间\n"
            report += "  - 睡前避免剧烈运动\n"
        else:
            report += "• 睡眠质量优秀，继续保持良好的睡眠习惯！\n"
        
        report += "\n" + "=" * 50
        return report


class RecordPage(QWidget):
    record_added = pyqtSignal()
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.records = []
        self.chart_window = None
        self.stats_window = None
        self.sleep_window = None
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        
        title_label = QLabel("心率变化记录")
        title_label.setFont(QFont("Arial", 18, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("color: #2C3E50; margin: 10px;")
        layout.addWidget(title_label)
        
        self.table = QTableWidget()
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(["序号", "心率 (BPM)", "时间"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                border: 1px solid #BDC3C7;
                border-radius: 5px;
                gridline-color: #ECF0F1;
            }
            QTableWidget::item {
                padding: 8px;
            }
            QTableWidget::item:selected {
                background-color: #3498DB;
                color: white;
            }
            QHeaderView::section {
                background-color: #34495E;
                color: white;
                padding: 8px;
                border: none;
                font-weight: bold;
            }
        """)
        layout.addWidget(self.table)
        
        button_layout = QHBoxLayout()
        
        chart_button = QPushButton("显示折线图")
        chart_button.setFont(QFont("Arial", 12))
        chart_button.setMinimumSize(120, 35)
        chart_button.setStyleSheet("""
            QPushButton {
                background-color: #3498DB;
                color: white;
                border: none;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #2980B9;
            }
        """)
        chart_button.clicked.connect(self.show_chart)
        button_layout.addWidget(chart_button)
        
        stats_button = QPushButton("统计面板")
        stats_button.setFont(QFont("Arial", 12))
        stats_button.setMinimumSize(100, 35)
        stats_button.setStyleSheet("""
            QPushButton {
                background-color: #9B59B6;
                color: white;
                border: none;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #8E44AD;
            }
        """)
        stats_button.clicked.connect(self.show_stats)
        button_layout.addWidget(stats_button)
        
        sleep_button = QPushButton("睡眠分析")
        sleep_button.setFont(QFont("Arial", 12))
        sleep_button.setMinimumSize(100, 35)
        sleep_button.setStyleSheet("""
            QPushButton {
                background-color: #27AE60;
                color: white;
                border: none;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #229954;
            }
        """)
        sleep_button.clicked.connect(self.show_sleep_analysis)
        button_layout.addWidget(sleep_button)
        
        button_layout.addStretch()
        
        clear_button = QPushButton("清空记录")
        clear_button.setFont(QFont("Arial", 12))
        clear_button.setMinimumSize(100, 35)
        clear_button.setStyleSheet("""
            QPushButton {
                background-color: #E74C3C;
                color: white;
                border: none;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #C0392B;
            }
        """)
        clear_button.clicked.connect(self.clear_records)
        button_layout.addWidget(clear_button)
        
        layout.addLayout(button_layout)

    def show_chart(self):
        if self.chart_window is None or not self.chart_window.isVisible():
            self.chart_window = ChartWindow(self)
            self.chart_window.setWindowTitle("心率折线图")
            self.chart_window.show()
        else:
            self.chart_window.raise_()
            self.chart_window.activateWindow()

    def show_stats(self):
        if self.stats_window is None or not self.stats_window.isVisible():
            self.stats_window = StatsWindow(self)
            self.stats_window.setWindowTitle("心率统计面板")
            self.stats_window.show()
        else:
            self.stats_window.raise_()
            self.stats_window.activateWindow()

    def show_sleep_analysis(self):
        if self.sleep_window is None or not self.sleep_window.isVisible():
            self.sleep_window = SleepAnalysisWindow(self)
            self.sleep_window.setWindowTitle("睡眠质量分析")
            self.sleep_window.show()
        else:
            self.sleep_window.raise_()
            self.sleep_window.activateWindow()

    def add_record(self, heart_rate, timestamp):
        record = HeartRateRecord(heart_rate, timestamp)
        self.records.append(record)
        
        row = self.table.rowCount()
        self.table.insertRow(row)
        
        self.table.setItem(row, 0, QTableWidgetItem(str(row + 1)))
        self.table.setItem(row, 1, QTableWidgetItem(str(heart_rate)))
        self.table.setItem(row, 2, QTableWidgetItem(timestamp.strftime("%Y-%m-%d %H:%M:%S")))
        
        self.table.scrollToBottom()
        self.record_added.emit()

    def clear_records(self):
        self.records.clear()
        self.table.setRowCount(0)


class HeartRateWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.show_hrv = False  # 默认关闭HRV
        self.obs_enabled = False  # 默认关闭OBS对接
        self.obs_file_path = "obs_heart_rate.txt"  # OBS数据文件路径
        self.current_date = date.today()  # 当前日期
        # 将默认数据存储路径改为软件的安装路径
        self.data_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "HeartRateData")  # 数据存储目录
        self.init_data_dir()  # 初始化数据目录
        self.init_ui()
        self.load_daily_records()  # 加载当日数据
        self.setup_daily_reset()  # 设置每日重置定时器

    def init_ui(self):
        self.setWindowTitle(f"心率广播接收器 {APP_VERSION}")
        self.setMinimumSize(500, 400)
        
        menubar = self.menuBar()
        menubar.setStyleSheet("""
            QMenuBar {
                background-color: #34495E;
                color: white;
                padding: 5px;
            }
            QMenuBar::item {
                background-color: transparent;
                padding: 5px 15px;
                border-radius: 3px;
            }
            QMenuBar::item:selected {
                background-color: #3498DB;
            }
            QMenu {
                background-color: #34495E;
                color: white;
                border: 1px solid #2C3E50;
            }
            QMenu::item {
                padding: 8px 25px;
            }
            QMenu::item:selected {
                background-color: #3498DB;
            }
        """)
        
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        
        nav_bar = QWidget()
        nav_bar.setFixedHeight(50)
        nav_bar.setStyleSheet("background-color: #2C3E50;")
        nav_layout = QHBoxLayout(nav_bar)
        nav_layout.setContentsMargins(20, 0, 20, 0)
        
        self.home_button = QPushButton("首页")
        self.home_button.setFont(QFont("Arial", 12))
        self.home_button.setCheckable(True)
        self.home_button.setChecked(True)
        self.home_button.setFixedHeight(40)
        self.home_button.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 0 20px;
            }
            QPushButton:hover {
                background-color: #34495E;
            }
            QPushButton:checked {
                background-color: #3498DB;
            }
        """)
        self.home_button.clicked.connect(lambda: self.switch_page(0))
        nav_layout.addWidget(self.home_button)
        
        self.record_button = QPushButton("记录")
        self.record_button.setFont(QFont("Arial", 12))
        self.record_button.setCheckable(True)
        self.record_button.setFixedHeight(40)
        self.record_button.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 0 20px;
            }
            QPushButton:hover {
                background-color: #34495E;
            }
            QPushButton:checked {
                background-color: #3498DB;
            }
        """)
        self.record_button.clicked.connect(lambda: self.switch_page(1))
        nav_layout.addWidget(self.record_button)
        
        nav_layout.addStretch()
        main_layout.addWidget(nav_bar)
        
        self.stack = QStackedWidget()
        
        self.home_page = HomePage()
        self.home_page.heart_rate_recorded.connect(self.on_heart_rate_recorded)
        self.home_page.update_hrv_visibility(self.show_hrv)  # 初始化HRV可见性
        self.set_obs_settings()  # 初始化OBS设置
        self.stack.addWidget(self.home_page)
        
        self.record_page = RecordPage()
        self.stack.addWidget(self.record_page)
        
        main_layout.addWidget(self.stack)
        
        settings_menu = menubar.addMenu("设置")
        
        # 显示动态实时卡路里选项
        self.calorie_action = QAction("显示动态实时卡路里", self)
        self.calorie_action.triggered.connect(self.show_calorie_settings)
        settings_menu.addAction(self.calorie_action)
        
        # 在设置菜单中添加HRV控制选项
        self.hrv_action = QAction("关闭心率变异性（HRV）" if self.show_hrv else "显示心率变异性（HRV）", self)
        self.hrv_action.triggered.connect(lambda: self.toggle_hrv())
        settings_menu.addAction(self.hrv_action)
        
        # 在设置菜单中添加OBS对接选项
        self.obs_action = QAction("对接OBS", self)
        self.obs_action.triggered.connect(self.show_obs_settings)
        settings_menu.addAction(self.obs_action)
        
        # 添加更改数据存储位置选项
        self.data_location_action = QAction("更改数据存储位置", self)
        self.data_location_action.triggered.connect(self.show_data_location_settings)
        settings_menu.addAction(self.data_location_action)
        
        settings_menu.addSeparator()
        
        update_action = QAction("检查更新", self)
        update_action.triggered.connect(self.check_update)
        settings_menu.addAction(update_action)
        
        data_menu = menubar.addMenu("数据")
        
        export_csv_action = QAction("导出为 CSV", self)
        export_csv_action.triggered.connect(self.export_to_csv)
        data_menu.addAction(export_csv_action)
        
        export_excel_action = QAction("导出为 Excel", self)
        export_excel_action.triggered.connect(self.export_to_excel)
        data_menu.addAction(export_excel_action)
        
        # 延迟执行更新检查，确保UI已加载
        from PyQt5.QtCore import QTimer
        QTimer.singleShot(1000, self.check_update_at_startup)

    def show_calorie_settings(self):
        dialog = CalorieSettingsDialog(self, self.home_page.get_calorie_settings())
        if dialog.exec_() == QDialog.Accepted:
            settings = dialog.get_settings()
            self.home_page.set_calorie_settings(settings)
            # 更新菜单项文本
            if settings['enabled']:
                self.calorie_action.setText("关闭动态实时卡路里")
            else:
                self.calorie_action.setText("显示动态实时卡路里")

    def show_obs_settings(self):
        """显示OBS对接设置对话框"""
        from PyQt5.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QLabel, QCheckBox, QPushButton, QFileDialog, QLineEdit, QGroupBox, QComboBox
        
        dialog = QDialog(self)
        dialog.setWindowTitle("OBS对接设置")
        dialog.setMinimumWidth(600)
        
        layout = QVBoxLayout(dialog)
        
        # 启用OBS对接选项
        self.obs_checkbox = QCheckBox("启用OBS对接")
        self.obs_checkbox.setChecked(self.obs_enabled)
        self.obs_checkbox.setStyleSheet("font-weight: bold; color: #2C3E50;")
        layout.addWidget(self.obs_checkbox)
        
        layout.addSpacing(10)
        
        # 输出方式选择
        output_group = QGroupBox("输出方式")
        output_layout = QVBoxLayout(output_group)
        
        self.output_type_combo = QComboBox()
        self.output_type_combo.addItem("文本文件 (简单数字)", "txt")
        self.output_type_combo.addItem("HTML文件 (带图片/动画)", "html")
        self.output_type_combo.setCurrentIndex(0 if not hasattr(self, 'obs_output_type') or self.obs_output_type == 'txt' else 1)
        self.output_type_combo.currentIndexChanged.connect(self.on_output_type_changed)
        output_layout.addWidget(QLabel("选择输出方式:"))
        output_layout.addWidget(self.output_type_combo)
        
        layout.addWidget(output_group)
        
        layout.addSpacing(10)
        
        # 文件路径设置
        file_group = QGroupBox("数据文件设置")
        file_layout = QVBoxLayout(file_group)
        
        path_layout = QHBoxLayout()
        path_label = QLabel("数据文件路径:")
        self.obs_path_edit = QLineEdit(self.obs_file_path)
        browse_button = QPushButton("浏览...")
        browse_button.clicked.connect(self.browse_obs_file)
        
        path_layout.addWidget(path_label)
        path_layout.addWidget(self.obs_path_edit, 1)
        path_layout.addWidget(browse_button)
        
        file_layout.addLayout(path_layout)
        
        # 图片设置
        image_layout = QHBoxLayout()
        image_label = QLabel("心率图标路径:")
        self.obs_image_edit = QLineEdit(getattr(self, 'obs_image_path', ""))
        image_browse_button = QPushButton("浏览...")
        image_browse_button.clicked.connect(self.browse_obs_image)
        
        image_layout.addWidget(image_label)
        image_layout.addWidget(self.obs_image_edit, 1)
        image_layout.addWidget(image_browse_button)
        
        file_layout.addLayout(image_layout)
        
        info_label = QLabel("OBS可以通过读取此文件来获取实时心率数据")
        info_label.setStyleSheet("color: #7F8C8D; font-size: 24px;")
        file_layout.addWidget(info_label)
        
        html_info_label = QLabel("使用HTML输出时，OBS需添加浏览器源并指向生成的HTML文件")
        html_info_label.setStyleSheet("color: #3498DB; font-size: 24px;")
        file_layout.addWidget(html_info_label)
        
        layout.addWidget(file_group)
        
        layout.addSpacing(20)
        
        # 使用说明部分
        usage_group = QGroupBox("使用说明")
        usage_layout = QVBoxLayout(usage_group)
        
        usage_steps = [
            "1. 启用OBS对接功能",
            "2. 选择输出方式（文本文件或HTML文件）",
            "3. 设置数据文件保存路径",
            "4. 选择自定义心率图标（仅HTML输出方式）",
            "5. 开始接收心率数据",
            "6. 在OBS中添加相应的源：",
            "   - 文本文件：添加文本源，指向生成的txt文件",
            "   - HTML文件：添加浏览器源，指向生成的HTML文件",
            "7. 调整源的大小和位置"
        ]
        
        for step in usage_steps:
            step_label = QLabel(step)
            step_label.setStyleSheet("color: #2C3E50; font-size: 16px;")
            usage_layout.addWidget(step_label)
        
        layout.addWidget(usage_group)
        
        layout.addSpacing(20)
        
        # 按钮布局
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        cancel_button = QPushButton("取消")
        cancel_button.setMinimumWidth(80)
        cancel_button.setStyleSheet("""
            QPushButton {
                background-color: #95A5A6;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px;
            }
            QPushButton:hover {
                background-color: #7F8C8D;
            }
        """)
        cancel_button.clicked.connect(dialog.reject)
        button_layout.addWidget(cancel_button)
        
        ok_button = QPushButton("确定")
        ok_button.setMinimumWidth(80)
        ok_button.setStyleSheet("""
            QPushButton {
                background-color: #27AE60;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px;
            }
            QPushButton:hover {
                background-color: #2ECC71;
            }
        """)
        ok_button.clicked.connect(lambda: self.save_obs_settings(dialog))
        button_layout.addWidget(ok_button)
        
        layout.addLayout(button_layout)
        
        dialog.exec_()

    def browse_obs_file(self):
        """浏览OBS数据文件路径"""
        # 使用对话框中当前选择的输出类型
        if hasattr(self, 'output_type_combo'):
            output_type = self.output_type_combo.currentData()
        else:
            output_type = getattr(self, 'obs_output_type', 'txt')
        
        # 更新初始文件名，确保后缀与输出类型匹配
        import os
        initial_file_path = self.obs_file_path
        if initial_file_path:
            base_name = os.path.splitext(initial_file_path)[0]
            initial_file_path = f"{base_name}.{output_type}"
        
        file_filter = "文本文件 (*.txt)" if output_type == 'txt' else "HTML文件 (*.html)"
        file_path, _ = QFileDialog.getSaveFileName(
            self, "保存OBS数据文件", 
            initial_file_path,
            file_filter
        )
        if file_path:
            # 确保文件后缀与选择的输出类型匹配
            base_name = os.path.splitext(file_path)[0]
            new_file_path = f"{base_name}.{output_type}"
            self.obs_path_edit.setText(new_file_path)

    def browse_obs_image(self):
        """浏览心率图标路径"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择心率图标", 
            "",
            "图片文件 (*.png *.jpg *.jpeg *.gif)"
        )
        if file_path:
            self.obs_image_edit.setText(file_path)

    def on_output_type_changed(self, index):
        """当输出类型改变时，更新文件路径扩展名"""
        output_type = self.output_type_combo.currentData()
        current_path = self.obs_path_edit.text()
        
        if current_path:
            # 移除旧的扩展名
            import os
            base_name = os.path.splitext(current_path)[0]
            # 添加新的扩展名
            new_path = f"{base_name}.{output_type}"
            self.obs_path_edit.setText(new_path)

    def save_obs_settings(self, dialog):
        """保存OBS设置"""
        self.obs_enabled = self.obs_checkbox.isChecked()
        self.obs_file_path = self.obs_path_edit.text()
        self.obs_output_type = self.output_type_combo.currentData()
        self.obs_image_path = self.obs_image_edit.text()
        
        # 更新菜单项文本
        if self.obs_enabled:
            self.obs_action.setText("关闭OBS对接")
        else:
            self.obs_action.setText("对接OBS")
        
        # 向HomePage传递OBS设置
        self.set_obs_settings()
        
        dialog.accept()

    def set_obs_settings(self):
        """向HomePage传递OBS设置"""
        self.home_page.obs_enabled = self.obs_enabled
        self.home_page.obs_file_path = self.obs_file_path
        self.home_page.obs_output_type = getattr(self, 'obs_output_type', 'txt')
        self.home_page.obs_image_path = getattr(self, 'obs_image_path', "")
        # 更新OBS对接网址显示
        self.home_page.update_obs_url_display()

    def version_compare(self, v1, v2):
        """比较两个版本号的大小
        返回 1 如果 v1 > v2
        返回 0 如果 v1 == v2
        返回 -1 如果 v1 < v2
        """
        # 移除版本号前缀的 'v'
        v1 = v1.lstrip('v')
        v2 = v2.lstrip('v')
        
        # 分割版本号为数字列表
        v1_parts = list(map(int, v1.split('.')))
        v2_parts = list(map(int, v2.split('.')))
        
        # 比较每个部分
        for i in range(max(len(v1_parts), len(v2_parts))):
            v1_part = v1_parts[i] if i < len(v1_parts) else 0
            v2_part = v2_parts[i] if i < len(v2_parts) else 0
            
            if v1_part > v2_part:
                return 1
            elif v1_part < v2_part:
                return -1
        
        return 0

    def check_update(self):
        try:
            req = urllib.request.Request(GITHUB_RELEASES_API)
            req.add_header('User-Agent', f'HeartRateReceiver/{APP_VERSION}')
            
            with urllib.request.urlopen(req, timeout=10) as response:
                data = json.loads(response.read().decode('utf-8'))
                latest_version = data.get('tag_name', '')
                release_url = data.get('html_url', '')
                release_notes = data.get('body', '')
            
            if not latest_version:
                QMessageBox.warning(self, "检查更新", "无法获取版本信息")
                return
            
            if self.version_compare(APP_VERSION, latest_version) >= 0:
                QMessageBox.information(
                    self, "检查更新",
                    f"当前已是最新版本\n\n当前版本: {APP_VERSION}"
                )
            else:
                msg = f"发现新版本!\n\n当前版本: {APP_VERSION}\n最新版本: {latest_version}"
                if release_notes:
                    msg += f"\n\n更新内容:\n{release_notes[:200]}"
                    if len(release_notes) > 200:
                        msg += "..."
                
                reply = QMessageBox.question(
                    self, "检查更新", msg,
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.Yes
                )
                
                if reply == QMessageBox.Yes:
                    # 尝试自动下载并安装新版本
                    self.download_and_install_update(data)
                    
        except urllib.error.HTTPError as e:
            if e.code == 404:
                QMessageBox.information(
                    self, "检查更新",
                    f"当前版本: {APP_VERSION}\n\n暂无发布版本\n\n请前往 GitHub 发布第一个版本"
                )
            else:
                QMessageBox.warning(self, "检查更新", f"HTTP错误: {e.code} {e.reason}")
        except urllib.error.URLError as e:
            QMessageBox.warning(self, "检查更新", f"网络错误: {str(e)}\n\n请检查网络连接")
        except Exception as e:
            QMessageBox.warning(self, "检查更新", f"检查更新失败: {str(e)}")

    def download_and_install_update(self, release_data):
        """自动下载并安装新版本"""
        try:
            # 显示下载进度对话框
            from PyQt5.QtWidgets import QProgressDialog
            progress = QProgressDialog("正在下载更新...", "取消", 0, 100, self)
            progress.setWindowTitle("下载更新")
            progress.setWindowModality(True)
            progress.show()
            
            # 获取下载链接
            assets = release_data.get('assets', [])
            download_url = None
            for asset in assets:
                if asset.get('name', '').endswith('.exe'):
                    download_url = asset.get('browser_download_url')
                    break
            
            if not download_url:
                QMessageBox.warning(self, "下载失败", "未找到可下载的安装文件")
                return
            
            # 下载文件
            import tempfile
            import os
            temp_dir = tempfile.gettempdir()
            file_name = os.path.join(temp_dir, f"HeartRateReceiver_{release_data.get('tag_name', 'latest')}.exe")
            
            def report_progress(count, block_size, total_size):
                percent = int(count * block_size * 100 / total_size)
                progress.setValue(percent)
                QApplication.processEvents()
                if progress.wasCanceled():
                    raise Exception("下载被用户取消")
            
            import urllib.request
            urllib.request.urlretrieve(download_url, file_name, reporthook=report_progress)
            
            progress.close()
            
            # 运行安装程序
            QMessageBox.information(self, "下载完成", "更新文件已下载完成，即将开始安装")
            
            import subprocess
            subprocess.Popen([file_name])
            
            # 退出当前应用程序
            QApplication.instance().quit()
            
        except Exception as e:
            QMessageBox.warning(self, "更新失败", f"自动更新失败: {str(e)}")
            # 如果自动更新失败，回退到打开浏览器
            release_url = release_data.get('html_url', '')
            if release_url:
                import webbrowser
                webbrowser.open(release_url)

    def check_update_at_startup(self):
        try:
            req = urllib.request.Request(GITHUB_RELEASES_API)
            req.add_header('User-Agent', f'HeartRateReceiver/{APP_VERSION}')
            
            with urllib.request.urlopen(req, timeout=10) as response:
                data = json.loads(response.read().decode('utf-8'))
                latest_version = data.get('tag_name', '')
                release_notes = data.get('body', '')
            
            if not latest_version:
                return
            
            if self.version_compare(APP_VERSION, latest_version) < 0:
                msg = f"发现新版本!\n\n当前版本: {APP_VERSION}\n最新版本: {latest_version}"
                if release_notes:
                    msg += f"\n\n更新内容:\n{release_notes[:200]}"
                    if len(release_notes) > 200:
                        msg += "..."
                
                msg += "\n\n⚠️  强制更新：请更新到最新版本后再使用。"
                
                reply = QMessageBox.critical(
                    self, "强制更新", msg,
                    QMessageBox.Ok | QMessageBox.Cancel,
                    QMessageBox.Ok
                )
                
                if reply == QMessageBox.Ok:
                    # 尝试自动下载并安装新版本
                    self.download_and_install_update(data)
                
                # 无论用户选择什么，都退出软件
                QApplication.instance().quit()
                
        except urllib.error.HTTPError as e:
            if e.code != 404:
                pass
        except urllib.error.URLError as e:
            pass
        except Exception as e:
            pass

    def init_data_dir(self):
        """初始化数据存储目录"""
        if not os.path.exists(self.data_dir):
            os.makedirs(self.data_dir)

    def setup_daily_reset(self):
        """设置每日重置定时器"""
        # 计算到今天24:00的时间差
        now = datetime.now()
        tomorrow = now.replace(hour=23, minute=59, second=59, microsecond=999999)
        time_to_reset = (tomorrow - now).total_seconds() * 1000
        
        # 设置定时器
        self.reset_timer = QTimer(self)
        self.reset_timer.setSingleShot(True)
        self.reset_timer.timeout.connect(self.daily_reset)
        self.reset_timer.start(int(time_to_reset))

    def daily_reset(self):
        """每日重置操作"""
        # 保存当日记录
        self.save_daily_records()
        
        # 清空记录面板
        self.record_page.clear_records()
        
        # 更新当前日期
        self.current_date = date.today()
        
        # 重新设置定时器
        self.setup_daily_reset()

    def save_daily_records(self):
        """保存当日记录到本地文件"""
        records = self.record_page.records
        if not records:
            return
        
        # 生成文件名，格式为"xx年xx月xx日"
        date_str = self.current_date.strftime("%Y年%m月%d日")
        file_path = os.path.join(self.data_dir, f"heart_rate_{date_str}.json")
        
        # 数据验证：确保所有记录都包含有效的心率值和时间戳
        valid_records = []
        for r in records:
            if hasattr(r, 'heart_rate') and hasattr(r, 'timestamp'):
                if isinstance(r.heart_rate, (int, float)) and r.heart_rate > 0:
                    valid_records.append(r)
        
        if not valid_records:
            print("没有有效的记录可保存")
            return
        
        # 优化数据格式，减少存储空间
        # 使用更紧凑的格式：[时间戳, 心率值]
        compact_records = [[r.timestamp.timestamp(), r.heart_rate] for r in valid_records]
        
        # 准备数据
        data = {
            "date": self.current_date.strftime("%Y-%m-%d"),
            "records": compact_records,
            "record_count": len(compact_records),
            "generated_at": datetime.now().isoformat()
        }
        
        # 保存数据，使用紧凑格式（无缩进）
        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, separators=(',', ':'))
        except Exception as e:
            print(f"保存数据失败: {str(e)}")

    def load_daily_records(self):
        """加载当日数据"""
        # 生成文件名，格式为"xx年xx月xx日"
        date_str = self.current_date.strftime("%Y年%m月%d日")
        file_path = os.path.join(self.data_dir, f"heart_rate_{date_str}.json")
        
        # 检查文件是否存在
        if not os.path.exists(file_path):
            return
        
        # 加载数据
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # 数据验证：确保数据结构完整
            if not isinstance(data, dict):
                print("数据格式无效：不是有效的JSON对象")
                return
            
            if 'records' not in data:
                print("数据格式无效：缺少records字段")
                return
            
            if not isinstance(data['records'], list):
                print("数据格式无效：records不是有效的列表")
                return
            
            # 清空当前记录
            self.record_page.clear_records()
            
            # 加载记录
            valid_count = 0
            for record_data in data.get('records', []):
                # 处理新的紧凑格式：[时间戳, 心率值]
                if isinstance(record_data, list) and len(record_data) == 2:
                    try:
                        timestamp = datetime.fromtimestamp(record_data[0])
                        heart_rate = record_data[1]
                        if isinstance(heart_rate, (int, float)) and heart_rate > 0:
                            self.record_page.add_record(heart_rate, timestamp)
                            valid_count += 1
                    except Exception as e:
                        print(f"解析记录失败: {str(e)}")
                # 兼容旧格式
                elif isinstance(record_data, dict):
                    try:
                        heart_rate = record_data.get('heart_rate')
                        timestamp_str = record_data.get('timestamp')
                        if isinstance(heart_rate, (int, float)) and heart_rate > 0 and timestamp_str:
                            timestamp = datetime.fromisoformat(timestamp_str)
                            self.record_page.add_record(heart_rate, timestamp)
                            valid_count += 1
                    except Exception as e:
                        print(f"解析记录失败: {str(e)}")
            
            print(f"成功加载 {valid_count} 条有效记录")
        except json.JSONDecodeError as e:
            print(f"JSON解析失败: {str(e)}")
        except Exception as e:
            print(f"加载数据失败: {str(e)}")

    def show_data_location_settings(self):
        """显示数据存储位置设置对话框"""
        from PyQt5.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QFileDialog, QLineEdit
        
        dialog = QDialog(self)
        dialog.setWindowTitle("数据存储位置设置")
        dialog.setMinimumWidth(500)
        
        layout = QVBoxLayout(dialog)
        
        # 当前存储位置
        current_label = QLabel(f"当前存储位置: {self.data_dir}")
        current_label.setWordWrap(True)
        current_label.setStyleSheet("color: #2C3E50;")
        layout.addWidget(current_label)
        
        layout.addSpacing(10)
        
        # 新存储位置
        path_layout = QHBoxLayout()
        path_label = QLabel("新存储位置:")
        self.new_data_dir_edit = QLineEdit(self.data_dir)
        browse_button = QPushButton("浏览...")
        browse_button.clicked.connect(self.browse_data_location)
        
        path_layout.addWidget(path_label)
        path_layout.addWidget(self.new_data_dir_edit, 1)
        path_layout.addWidget(browse_button)
        
        layout.addLayout(path_layout)
        
        # 提示信息
        hint_label = QLabel("提示：系统会在选择的位置自动创建'HeartRateData'文件夹来存储数据")
        hint_label.setWordWrap(True)
        hint_label.setStyleSheet("color: #7F8C8D; font-style: italic;")
        layout.addWidget(hint_label)
        
        layout.addSpacing(20)
        
        # 按钮布局
        button_layout = QHBoxLayout()
        cancel_button = QPushButton("取消")
        cancel_button.clicked.connect(dialog.reject)
        
        ok_button = QPushButton("确定")
        ok_button.clicked.connect(dialog.accept)
        
        button_layout.addStretch()
        button_layout.addWidget(cancel_button)
        button_layout.addWidget(ok_button)
        
        layout.addLayout(button_layout)
        
        if dialog.exec_() == QDialog.Accepted:
            new_dir = self.new_data_dir_edit.text().strip()
            if new_dir and new_dir != self.data_dir:
                try:
                    # 在选择的位置创建HeartRateData文件夹
                    data_dir = os.path.join(new_dir, "HeartRateData")
                    
                    # 确保目录存在
                    if not os.path.exists(data_dir):
                        os.makedirs(data_dir)
                    
                    # 更新数据目录
                    old_dir = self.data_dir
                    self.data_dir = data_dir
                    
                    # 立即保存当前记录到新位置
                    self.save_daily_records()
                    
                    QMessageBox.information(self, "成功", f"数据存储位置已更新，当前记录已保存到新位置: {data_dir}")
                except Exception as e:
                    # 恢复到原目录
                    self.data_dir = old_dir
                    QMessageBox.critical(self, "错误", f"设置数据存储位置失败: {str(e)}")

    def browse_data_location(self):
        """浏览数据存储位置"""
        directory = QFileDialog.getExistingDirectory(
            self, "选择数据存储目录", self.data_dir
        )
        if directory:
            self.new_data_dir_edit.setText(directory)

    def export_to_csv(self):
        records = self.record_page.records
        if not records:
            QMessageBox.warning(self, "导出数据", "没有数据可导出")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "导出为 CSV", 
            f"心率记录_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            "CSV文件 (*.csv)"
        )
        
        if not file_path:
            return
        
        try:
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(['序号', '心率 (BPM)', '时间'])
                for i, record in enumerate(records, 1):
                    writer.writerow([
                        i,
                        record.heart_rate,
                        record.timestamp.strftime('%Y-%m-%d %H:%M:%S')
                    ])
            
            QMessageBox.information(self, "导出成功", f"数据已导出到:\n{file_path}")
        except Exception as e:
            QMessageBox.warning(self, "导出失败", f"导出失败: {str(e)}")

    def export_to_excel(self):
        records = self.record_page.records
        if not records:
            QMessageBox.warning(self, "导出数据", "没有数据可导出")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "导出为 Excel",
            f"心率记录_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel文件 (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            wb = Workbook()
            ws = wb.active
            ws.title = "心率记录"
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            headers = ['序号', '心率 (BPM)', '时间']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            for i, record in enumerate(records, 1):
                ws.cell(row=i+1, column=1, value=i).border = thin_border
                ws.cell(row=i+1, column=2, value=record.heart_rate).border = thin_border
                ws.cell(row=i+1, column=3, value=record.timestamp.strftime('%Y-%m-%d %H:%M:%S')).border = thin_border
            
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 25
            
            wb.save(file_path)
            
            QMessageBox.information(self, "导出成功", f"数据已导出到:\n{file_path}")
        except ImportError:
            QMessageBox.warning(self, "导出失败", "请先安装 openpyxl 库:\npip install openpyxl")
        except Exception as e:
            QMessageBox.warning(self, "导出失败", f"导出失败: {str(e)}")

    def switch_page(self, index):
        self.stack.setCurrentIndex(index)
        self.home_button.setChecked(index == 0)
        self.record_button.setChecked(index == 1)

    def on_heart_rate_recorded(self, heart_rate, timestamp):
        self.record_page.add_record(heart_rate, timestamp)
        # 实时保存数据
        self.save_daily_records()

    def toggle_hrv(self):
        # 切换HRV状态
        self.show_hrv = not self.show_hrv
        self.home_page.update_hrv_visibility(self.show_hrv)
        if self.show_hrv:
            self.hrv_action.setText("关闭心率变异性（HRV）")
        else:
            self.hrv_action.setText("显示心率变异性（HRV）")

    def closeEvent(self, event):
        self.home_page.cleanup()
        event.accept()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = HeartRateWindow()
    window.show()
    sys.exit(app.exec_())
